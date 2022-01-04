from openpyxl import load_workbook

class Utils:

    def read_data_from_excel(filename, sheet):
        datalist = []
        wb = load_workbook(filename=filename)
        sh = wb[sheet]
        row_num = sh.max_row
        col_num = sh.max_column
        for i in range(2, row_num+1):
            row = []
            for j in range(1, col_num+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist